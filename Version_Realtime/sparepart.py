import cv2
import numpy
import openpyxl
from openpyxl.styles import Font
import subprocess
import sqlite3 


# sqlite
def insert_data_oke(line,part,quantity,total_box,capacity_hour,total_pcs,cycle_time,target,actual,loading_time,name_hikitori,data_date,data_time):
    conn = sqlite3.connect('/home/otics/on/project_pt_otics_ai_hla/core_engine/hla.db')
    cursor = conn.cursor()
    data = (line,part,quantity,total_box,capacity_hour,total_pcs,cycle_time,target,actual,loading_time,name_hikitori,data_date,data_time)
    task = "INSERT INTO data_production_oke (line,part,quantity,total_box,capacity_hour,total_pcs,cycle_time,target,actual,loading_time,name_hikitori,data_date,data_time) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)"
    cursor.execute(task, data)
    conn.commit()
    conn.close()
def insert_data_ng(line,part,quantity,total_box,capacity_hour,total_pcs,cycle_time,target,actual,loading_time,name_hikitori,data_date,data_time):
    conn = sqlite3.connect('/home/otics/on/project_pt_otics_ai_hla/core_engine/hla.db')
    cursor = conn.cursor()
    data = (line,part,quantity,total_box,capacity_hour,total_pcs,cycle_time,target,actual,loading_time,name_hikitori,data_date,data_time)
    task = "INSERT INTO data_production_ng (line,part,quantity,total_box,capacity_hour,total_pcs,cycle_time,target,actual,loading_time,name_hikitori,data_date,data_time) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)"
    cursor.execute(task, data)
    conn.commit()
    conn.close()

# opencv
def put_text_on_frame(frame, text, position, font_scale, font_thickness, box_size):
    (text_width, text_height), baseline = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, font_scale, font_thickness)
    box_width = text_width + 20
    box_height = text_height + 20
    box = numpy.ones((box_height, box_width, 3), dtype=numpy.uint8) * 255
    cv2.putText(box, text, (10, box_height - 10), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 0, 0), font_thickness, cv2.LINE_AA)
    x, y = position
    frame[y:y + box_height, x:x + box_width] = box
    cv2.rectangle(frame, (x, y), (x + box_width, y + box_height), (56, 10, 0), 1)

#exel
def make_report(nama_file, data):
    try:
        workbook = openpyxl.load_workbook(nama_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Laporan Project HLA'
        sheet['A1'].font = Font(size=16, bold=True)
        header = ['Tanggal', 'Part', 'Jumlah', 'Delay', 'Total']
        for col_num, col_title in enumerate(header, 1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = col_title
            cell.font = Font(bold=True)
    row_num = sheet.max_row + 1
    for col_num, col_value in enumerate(data, 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.value = col_value
    workbook.save(nama_file)
#os
def shutdown():
        subprocess.run(["shutdown", "/s", "/t", "0"])

